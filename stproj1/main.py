import streamlit as st
import yaml
import bcrypt
import streamlit_authenticator as stauth
import pandas as pd
from openpyxl import load_workbook
import os
import requests  # Make sure to import requests

# Initialize session state
if 'in_session' not in st.session_state:
    st.session_state.in_session = False

# Main landing page
def landing():
    st.title("🚀 Streamlit Application")
    st.sidebar.header("Choose an Operation")
    option = st.sidebar.radio('Select Role', ['Admin Login', 'User Login', 'User Signup', 'IP Webcam Stream'])

    if option == 'Admin Login':
        admin_login()
    elif option == 'User Login':
        user_login()
    elif option == 'User Signup':
        signup()
    elif option == 'IP Webcam Stream':
        ip_webcam_stream()

# Load credentials from YAML
def load_credentials():
    base_path = os.path.dirname(__file__)  
    file_path = os.path.join(base_path, 'config.yaml')  
    
    if not os.path.exists(file_path):
        st.error(f"Config file not found at {file_path}. Please check the file path.")
        return None
    
    with open(file_path, 'r') as file:
        config = yaml.safe_load(file)
    return config

# Save new user credentials to YAML
def save_credentials(new_user):
    base_path = os.path.dirname(__file__)
    file_path = os.path.join(base_path, 'config.yaml')

    if not os.path.exists(file_path):
        st.error(f"Config file not found at {file_path}. Cannot save new user.")
        return
    
    with open(file_path, 'r') as file:
        config = yaml.safe_load(file)

    for username, user_info in new_user.items():
        config['credentials']['usernames'][username] = {
            'mail': user_info['mail'],
            'password': user_info['password'],
            'role': user_info['role'],
            'name': username  
        }

    with open(file_path, 'w') as file:
        yaml.dump(config, file)

# Save data to Excel
def save_data_to_excel(name, choice):
    base_path = os.path.dirname(__file__)
    file_name = os.path.join(base_path, 'form_data.xlsx')

    data = pd.DataFrame([[name, choice]], columns=['Name', 'Gender'])

    if not os.path.exists(file_name):
        data.to_excel(file_name, index=False)
    else:
        try:
            existing_data = pd.read_excel(file_name)
            duplicate = (existing_data['Name'] == name) & (existing_data['Gender'] == choice)

            if duplicate.any():
                st.warning(f"Duplicate entry found: {name}, {choice}")
            else:
                with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    book = load_workbook(file_name)
                    sheet = book.active
                    startrow = sheet.max_row
                    data.to_excel(writer, index=False, header=False, startrow=startrow)
                    st.success(f"Data saved: {name}, {choice}")
        except Exception as e:
            st.error(f"Error: {e}")

# Admin login
def admin_login():
    st.sidebar.subheader("Admin Login")

    config = load_credentials()
    if config is None:
        return  

    authenticator = stauth.Authenticate(
        config['credentials'],
        config['cookie']['name'],
        config['cookie']['key'],
        config['cookie']['expiry_days']
    )

    name, auth_status, email = authenticator.login(location='sidebar', key='admin_login')

    if auth_status:
        user_role = config['credentials']['usernames'][email]['role']
        if user_role == 'admin':
            admin_dashboard()
        else:
            st.error("Access denied: Not an admin.")
    elif auth_status is False:
        st.error("ERROR: Invalid credentials.")
    elif auth_status is None:
        st.warning("Please enter your email and password.")

# User login
def user_login():
    st.sidebar.subheader("User Login")

    config = load_credentials()
    if config is None:
        return  

    authenticator = stauth.Authenticate(
        config['credentials'],
        config['cookie']['name'],
        config['cookie']['key'],
        config['cookie']['expiry_days']
    )

    username, auth_status, email = authenticator.login(location='main', key='user_login')

    if auth_status:
        if username in config['credentials']['usernames']:
            user_role = config['credentials']['usernames'][username]['role']
            if user_role == 'user':
                st.session_state['username'] = username  
                dashboard()
            else:
                st.error("Access denied.")
        else:
            st.error("User not found. Please sign up first.")
    elif auth_status is False:
        st.error("ERROR: Invalid credentials.")
    elif auth_status is None:
        st.warning("Please enter your email and password.")

# Admin dashboard
def admin_dashboard():
    st.success("Login successful! Welcome, Admin.")
    st.title('Data Collection Form')

    with st.form(key='data_form'):
        name = st.text_input('Enter your name')
        choice = st.radio('Gender:', ['Male', 'Female', 'Other'])

        submit_button = st.form_submit_button(label='Submit')

    if st.button('Load Data'):
        base_path = os.path.dirname(__file__)
        file_name = os.path.join(base_path, 'form_data.xlsx')

        if os.path.exists(file_name):
            df = pd.read_excel(file_name)
            st.dataframe(df)
        else:
            st.warning("No data available to show, fill the form first")

    if submit_button:
        if name and choice:
            save_data_to_excel(name, choice)
        else:
            st.error("Please enter your name.")

# User dashboard
def dashboard():
    st.success("Login successful! Welcome, User.")
    st.header("You have reading permissions only.")
    if st.button('Load Data'):
        base_path = os.path.dirname(__file__)
        file_name = os.path.join(base_path, 'form_data.xlsx')

        if os.path.exists(file_name):
            df = pd.read_excel(file_name)
            st.dataframe(df)
        else:
            st.warning("No data available to show, fill the form first.")

# User signup
def signup():
    st.sidebar.subheader("User Signup")

    name = st.text_input("Username")
    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    if st.button("Sign Up"):
        if email and password and name:
            users = load_credentials()
            if users is None:
                return  

            if name in users['credentials']['usernames']:
                st.error("Username already exists!")
            else:
                salt = bcrypt.gensalt()
                hashed_pw = bcrypt.hashpw(password.encode('utf-8'), salt)
                new_user = {
                    name: {
                        'name': name,
                        'mail': email,
                        'password': hashed_pw.decode('utf-8'),
                        'role': 'user'
                    }
                }

                save_credentials(new_user)
                st.success("Registration Successful. You can now login.")
        else:
            st.error("Please enter all the details.")

# IP Webcam Stream Function
def ip_webcam_stream():
    st.title("IP Webcam Stream")

    ip_address = "192.168.29.156"  
    camera_url = f"http://{ip_address}:8080/video"

    if st.button('Start Stream'):
        st.write("Fetching video from mobile...")

        try:
            video_feed = requests.get(camera_url, stream=True)

            if video_feed.status_code == 200:
                st.video(camera_url)
            else:
                st.error(f"Error fetching video stream. Status code: {video_feed.status_code}")
        except Exception as e:
            st.error(f"An error occurred: {e}")

# Main execution
if __name__ == "__main__":
    landing()
